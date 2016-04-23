# -*- coding:utf-8 -*-
from openpyxl import load_workbook
import redis
import random
import json

data = load_workbook('sensitive_words.xlsx')
table = data.get_sheet_by_name('Sheet2')
category = ['政治', '军事', '法律', '意识形态', '民运']
#print table.cell('A1').value
r = redis.StrictRedis(host='219.224.135.97', port='6380', db=15)
for i in range(1,549):
    word = table.cell(row=i, column=0).value
    level = table.cell(row=i, column=1).value
    index = random.randint(0,4)
    r.hset('sensitive_words', word, json.dumps([level, category[index]]))

#r.hset('recommend_sensitive_words_20130901', '洪秀柱', json.dumps([['1093622153', '3270699555'], 3]))
#r.hset('recommend_sensitive_words_20130901', '港灿', json.dumps([['1686546714'], 1]))
#r.hset('recommend_sensitive_words_20130901', '钟屿晨', json.dumps([['1887344341', '3183040584'], 2]))


